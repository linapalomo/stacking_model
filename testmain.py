import openpyxl
import subprocess
from datetime import datetime

num_runs = 5  # Replace with your desired number of runs
keyword = "Result -"  # Replace with the specific keyword or pattern to match

# Generate a timestamp or unique identifier for the file name
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS

# Create a new Excel file with the unique file name
file_name = f"output_{timestamp}.xlsx"
workbook = openpyxl.Workbook()
worksheet = workbook.active
    
for i in range(num_runs):
    result0 = subprocess.run(["python", "ztest0.py"], capture_output=True, text=True)
    output0 = result0.stdout
    
    
    result1 = subprocess.run(["python", "ztest1.py"], capture_output=True, text=True)
    output1 = result1.stdout
    
    result2 = subprocess.run(["python", "ztest2.py"], capture_output=True, text=True)
    output2 = result2.stdout
    
    # Find the start line for the subprocess outputs
    lines1 = output1.split("\n")
    start_line1 = next((index for index, line in enumerate(lines1) if keyword in line), None)

    lines2 = output2.split("\n")
    start_line2 = next((index for index, line in enumerate(lines2) if keyword in line), None)


    if start_line1 is not None and start_line2 is not None: 
        relevant_output1 = "\n".join(lines1[start_line1:])
        relevant_output2 = "\n".join(lines2[start_line2:])
        
    else:
        relevant_output1 = ""
        relevant_output2 = ""
        

    # Add a marker or label for the new run
    marker = f"Run {i+1}:\n"

    # Add the marker and relevant output data to the Excel file
    row = i+1
    worksheet.cell(row=row, column=1, value=marker + relevant_output1)
    worksheet.cell(row=row, column=2, value=marker + relevant_output2)
    

# Save the Excel file with the unique file name
workbook.save(file_name)



