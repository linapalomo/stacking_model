# MAIN FILE. 
#THIS FILE RUNS ALL THE THREE FILES OF THE SIMULATION.
#1. Runs the distribution file 0 that generates the random landscape.
#2. runs the first offsetting scheme with fixed land use trading.
#3. runs the second offsetting scheme with fixed ESS trading.
#4. saves the results for every combination in an excel file. 

import subprocess
import openpyxl
from datetime import datetime

num_runs = 50  # Replace with number of runs
keyword = "Result - "  # specific keyword or pattern to match

# Generate a timestamp or unique identifier for the file name
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS

# Create a new Excel file with the unique file name
file_name = f"output63_nuevo_50final_{timestamp}.xlsx"
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define the specific data to save from each file
data_to_save = {
    "stack_0_distr.py": ["Current costs:","Current Land uses:", "Total ES1:", "Total ES2:","mean_lu1:", "mean_lu2:","mean benefitLU1","mean benefitLU2:","Proportion LU1:", "Proportion LU2:" ],
    "stack_1_lu.py": ["Objective value:","Optimized Land Use distribution with LU restrictions:", "Optimized total ES1:", "Optimized total ES2:", "mean_lu1:","mean_lu2:", "mean benefitLU1:", "mean benefitLU2:","Proportion LU1:","Proportion LU2:","total LU0:","total LU1:", "total LU2:" ],
    "stack_2_ess.py": ["Objective value:","Optimized Land Use with benefits restriction:", "Optimized total benefit1/ES1:", "Optimized total benefit2/ES2:", "tot opti LU0:","tot opti LU1:","tot opti LU2:"]
}


'''row = 1
col = 1
for file, data_fields in data_to_save.items():
    worksheet.cell(row=row, column=col, value=file)
    for idx, field in enumerate(data_fields, start=1):
        worksheet.cell(row=row+idx, column=col, value=field)
    col += 2'''

for i in range(num_runs):
    row = 1

    for file, data_fields in data_to_save.items():
        result = subprocess.run(["python", file], capture_output=True, text=True)
        output = result.stdout
        
        relevant_output = []

        for line in output.split("\n"):
            for field in data_fields:
                if field in line:
                    "Current costs:","Current Land uses:", "Total ES1:", "Total ES2:", 
                    "Objective value:","Optimized Land Use distribution with LU restrictions:", "Optimized total ES1:", "Optimized total ES2:"
                    if "exclude_this_text" not in line:
                        relevant_output.append(line)
                    break
                
        """if file == "stack_0_distr.py" and i == 0:  # Add additional outputs only in the first run of ztest0
            additional_outputs = ["mean_lu1:","mean_lu2:", "mean benefitLU1:", "mean benefitLU2:","Proportion LU1:","Proportion LU2:"]
            if "mean_lu1:" in relevant_output:
                meanlu1 = float(relevant_output[relevant_output.index("mean_lu1") + 1])
                additional_outputs.append(f"mean_lu1 {meanlu1}")

            if "mean_lu2" in relevant_output:
                meanlu2 = float(relevant_output[relevant_output.index("mean_lu2") + 1])
                additional_outputs.append(f"mean_lu2 {meanlu2}")

            if "mean benefitLU1:" in relevant_output:
                mean_be1 = float(relevant_output[relevant_output.index("mean benefitLU1:") + 1])
                additional_outputs.append(f"mean benefitLU1: {mean_be1}")

            if "mean benefitLU2:" in relevant_output:
                mean_be2 = float(relevant_output[relevant_output.index("mean benefitLU2:") + 1])
                additional_outputs.append(f"mean benefitLU2: {mean_be2}")
                
            if "Proportion LU1:" in relevant_output:
                prop_lu1 = float(relevant_output[relevant_output.index("Proportion LU1:") + 1])
                additional_outputs.append(f"mean benefitLU2: {prop_lu1}")
                
            if "Proportion LU2:" in relevant_output:
                prop_lu2 = float(relevant_output[relevant_output.index("Proportion LU2:") + 1])
                additional_outputs.append(f"mean benefitLU2: {prop_lu2}")

            relevant_output.extend(additional_outputs)
"""
        row += len(data_fields) + 1
        for idx, value in enumerate(relevant_output, start=1):
            worksheet.cell(row=row+idx, column=(i*2)+2, value=value)

# Save the Excel file with the unique file name
workbook.save(file_name)
